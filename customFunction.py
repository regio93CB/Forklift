import pandas as pd
from openpyxl import load_workbook
import time, struct, cv2, os
import subprocess as sp
from termcolor import colored
from datetime import datetime
from pyzbar.pyzbar import decode
import re, threading, numpy as np
from collections import Counter
import matplotlib.pyplot as plt
from PIL import Image, ExifTags
import copy, json


class Simbolo:
    def __init__(self, nome, componente):
        self.nome = nome
        self.componente = componente
        
    def nome(self):
        return self._nome
    
    def componente(self):
        return self._componente
    
    def __str__(self):
        return "{} - {}".format(self.nome, self.componente)

class MyTimer(threading.Thread):

        def __init__(self):
            super().__init__()
            self.stop_event = threading.Event()
            self.paused = threading.Event()
            self.paused.set()
            self.time = 0

        def run(self):

            while not self.stop_event.is_set():
                self.paused.wait()
                self.time += 1
                time.sleep(1)
                # if not self.time  == 0: print("Counter", self.time)

        def resume(self):
            self.paused.set()
    
        def pause(self):
            self.paused.clear()

        def stop(self):
            self.stop_event.set()
            self.time = 0



def shutdown():  os.system('echo "babilonia" | sudo -S shutdown -h now')
def reboot():    os.system('echo "babilonia" | sudo -S reboot')   

def saveFIXImage(imgName, im0, folder_path):
    
    # Creare un'immagine nera

    imgPath = folder_path + imgName +".jpg"

    # Salva l'immagine nera
    cv2.imwrite(imgPath, im0)
        
#calcolo il tempo attuale 
def calculate_time():
    
    current_time = time.localtime()
    
    current_h = current_time.tm_hour 
    current_m = current_time.tm_min
    current_s = current_time.tm_sec
    
    time_now = f"{current_h:02d}:{current_m:02d}:{current_s:02d} sec"
    
    return time_now
    
#aggiusto i path dentro l'excel 
def adjust_path(tipo_pezzo): 
    
    # print("Tipo di pezzo da vedere: ", tipo_pezzo)
     
    if tipo_pezzo == "attacco_leva": 
        tipo_pezzo = "attacchi_leva"
        
    elif tipo_pezzo == "ghiera_1" or tipo_pezzo == "ghiera_2" or tipo_pezzo == "ghiera_3": 
        tipo_pezzo = "ghiere"
        
    elif tipo_pezzo == "distanziale_ghiera_1" or tipo_pezzo == "distanziale_ghiera_2": 
        tipo_pezzo = "distanziali_ghiera"
        
    elif tipo_pezzo == "semipomello": 
        tipo_pezzo = "semipomelli"
        
    elif tipo_pezzo == "pulsante": 
        tipo_pezzo = "pulsanti"
    else:
        print(tipo_pezzo, "Nome componente sconosciuto")
        exit()
        
    return tipo_pezzo

def leggi_dati_json(file_path):
    """Legge il file JSON e restituisce i dati come dizionario."""
    with open(file_path, 'r') as file:
        return json.load(file)


def getCropValuesJson(file_path):
    """Carica i valori di crop dal file JSON."""
    dati = leggi_dati_json(file_path)
    crop_vars = {}
    for key, value in dati.get("crop_values", {}).items():
        crop_vars[f'crop_xi{key}'] = value[0]
        crop_vars[f'crop_yi{key}'] = value[1]
        crop_vars[f'crop_xf{key}'] = value[2]
        crop_vars[f'crop_yf{key}'] = value[3]
    return crop_vars

#Estrai la matrice da yoo 
def yolo2Matrix(coordinates_list, distance_tollerance):
    
    coordinates_list = remove_concetric_symbols(coordinates_list, 0.03)
            
    coordinates_list.sort(key=lambda x: (x[2], -x[3]))  # Ordina per X crescente e Y decrescente
    
    # for row in coordinate_list: print(f"|{row[0]:<14}| |{row[2]:.3f}| |{row[3]:<.3f}|")
        
    result = []
    columns = []  # Ora usiamo una lista di colonne con il loro valore Y

    for symbol in coordinates_list:
        added_to_existing_column = False
        
        for column in columns:
            last_symbol_in_column = column[-1]
            if abs(last_symbol_in_column[2] - symbol[2]) <= float(distance_tollerance):
                column.append(symbol)
                added_to_existing_column = True
                break
        
        if not added_to_existing_column: columns.append([symbol])

    for column in columns: column.sort(key=lambda x: -x[3])  # Ordina la colonna in base al valore Y
    for column in columns: result.append([symbol[0] for symbol in column])

    return result    

#TODO: A CASA PROVIAMO TUTTO 
def newyolo2Matrix(coordinates_list, distance_tollerance):
    
    coordinates_list.sort(key=lambda x: (x[2], -x[3]))  # Ordina per X crescente e Y decrescente
    
    # for row in coordinate_list: print(f"|{row[0]:<14}| |{row[2]:.3f}| |{row[3]:<.3f}|")

    print("Sorted Coordinate List")
    for i in coordinates_list: print(i)
    
    result = []
    columns = []  # Ora usiamo una lista di colonne con il loro valore Y

    for symbol in coordinates_list:
        added_to_existing_column = False
        
        
        for column in columns:
            last_symbol_in_column = column[-1]
            
            if abs(last_symbol_in_column[2] - symbol[2]) <= float(distance_tollerance):
                
                column.append(symbol)
                added_to_existing_column = True
                break
        
        if not added_to_existing_column: columns.append([symbol])

    for column in columns: column.sort(key=lambda x: -x[3])  # Ordina la colonna in base al valore Y
    for column in columns: result.append([symbol[0] for symbol in column])

    return result    

#carico i valori di crop da un file txt già esistente
def getCropValuesTxt(txt_path):
    
    def leggi_file(file_path):
        
        dati = []
        with open(file_path, 'r') as file:
            for line in file:
                dati.append(list(map(int, line.strip().split(', '))))
        return dati
    
    dati = leggi_file(txt_path)
    crop_vars = {}
    for riga in dati:
        numero = riga[0]
        crop_vars[f'crop_xi{numero}'] = riga[1]
        crop_vars[f'crop_yi{numero}'] = riga[2]
        crop_vars[f'crop_xf{numero}'] = riga[3]
        crop_vars[f'crop_yf{numero}'] = riga[4]
    return crop_vars 

#estrai la matrice di classi dalla vbom
def vbom2Matrix(input_codice, vbom_path):

    matrix = []
    index_column = 0
    the_last_was_empty = False

    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value == input_codice: 
                
                riga = cell.row 
                col = cell.column
                
                for i in range(2,9):
                    
                    codice_componente = sheet.cell(row=riga, column = i).value
                    
                    if codice_componente is not None:
                        
                        nome_componente = sheet.cell(row=1, column = i).value
                        
                        # Trova il tipo di pezzo nella prima riga aggiustando il nome 
                        tipo_pezzo = adjust_path(nome_componente)
                                                
                        # Vai al foglio corispondente al tipo di pezzo
                        foglio_tipo_pezzo = wb[tipo_pezzo]                         
                                                                        
                        #andiamo a cercare i simboli sopra il componente
                        for row in foglio_tipo_pezzo.iter_rows(min_row=2, max_row=foglio_tipo_pezzo.max_row, min_col=1, max_col=1):
    
                            for cell in row:
                                
                                if cell.value == codice_componente:
                                    
                                    #excel parte da 1 e no da 0 
                                    colonna_iniziale = cell.column + 1
                                    colonna_finale = foglio_tipo_pezzo.max_column +1
                                    
                                    for col in range(colonna_iniziale, colonna_finale):
                                        
                                        cell_corrente = foglio_tipo_pezzo.cell(row=cell.row, column=col) 
                                        
                                        # print(f"cell corrente: {cell_corrente.value }")
                                        # print(f"Simbolo: {cell_corrente.value}, TLWE: {the_last_was_empty}" )
                                        
                                        if cell_corrente.value is not None and cell_corrente.value != "":
                                            
                                            symbol = Simbolo(nome = f"{cell_corrente.value}", componente= f"{nome_componente}")
                                            
                                            if the_last_was_empty == True:
                                                
                                                index_column += 1
                                                matrix.insert(index_column, [symbol])
                                                the_last_was_empty = False
                                               
                                            elif matrix == []:
                                                
                                                matrix = [[symbol]]
                                                
                                            else:
                                                matrix[index_column].append(symbol)
                                                the_last_was_empty = False

                                        elif cell_corrente.value is None or cell_corrente.value != "":
                                            
                                            the_last_was_empty = True
                                        
    return matrix

def findButtonSymbol(vbom_path, input_codice):
    
    symbol = None
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    
    #Entro nel foglio dei modelli devio e cerco il codice del pomello
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value == input_codice: 
                
                riga = cell.row 
                col = cell.column
                    
                codice_pulsante = sheet.cell(row=riga, column = 9).value
                                 
                # Vai al foglio corispondente al tipo di pezzo
                foglio_tipo_pezzo = wb["pulsanti"]                         
                                                                
                #andiamo a cercare i simboli sopra il componente
                for row in foglio_tipo_pezzo.iter_rows(min_row=2, max_row=foglio_tipo_pezzo.max_row, min_col=1, max_col=1):

                    for cell in row:
                        
                        if cell.value == codice_pulsante:
                            
                            #excel parte da 1 e no da 0 
                            colonna_iniziale = cell.column + 1
                            colonna_finale = foglio_tipo_pezzo.max_column +1
                            
                            for col in range(colonna_iniziale, colonna_finale):
                                
                                cell_corrente = foglio_tipo_pezzo.cell(row=cell.row, column=col) 
                                
                                if cell_corrente.value is not None and cell_corrente.value != "": symbol = cell_corrente.value
    
    return symbol
                                    
#converti la matrice di classi in una matrice normale
def class_vbom2vbom(class_vbom):
    
    vbom = []
    column = []
    
    #converto la matrice di oggetti in matrice di stringhe per il confronto
    for class_column in class_vbom:
        
        for class_symbol in class_column:
            
            symbol = class_symbol.nome 
            
            if column is None: column = [[symbol]]
            else: column.append(symbol)
        
        if vbom is None: vbom = [column]
        else: vbom.append(column)
    
        column = []        
    
    return vbom

#trova se esiste un componente con quelle caratteristiche
def findComponent(symbols, vbom_path):
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)

    component_row = []
                        
    for tipo_pezzo in wb.sheetnames[2:]:
        
        foglio_tipo_pezzo = wb[tipo_pezzo]
                                                        
        #andiamo a cercare i simboli sopra il componente
        for row in foglio_tipo_pezzo.iter_rows(min_row=2, max_row=foglio_tipo_pezzo.max_row, min_col=2, max_col=7):
            
            component_row = []

            for cell in row:
                
                if cell.value != None:
                    
                    if component_row == []: component_row = [cell.value]
                    else: component_row += [cell.value]
                    
            # if component_row != []: print(f"{codice_componente} nella riga {cell.row} per il componente {tipo_pezzo} ha la cella {component_row}")
            if component_row == symbols: print(f"Simboli appartenenti al componente {tipo_pezzo} {foglio_tipo_pezzo.cell(row=cell.row, column = 1).value}")


def to_scalar(value):
    """Converte un valore in float, gestendo sia Tensor che tipi numerici."""
    return value.item() if hasattr(value, 'item') else value

def new_matrices_match(class_vbom, vbom, yolo, frontal_check, fixture_check, frontal_yolo, frontal_vbom, good_counter, bad_counter, status_vector, is_alligned):
    
    """Errore sempre: "Errore rilevato"

    Returns:
    
        1 - Attacco leva 
        2 - Ghiera
        3 - Distanziale 
        4 - Semipomello 
        5 - Pulsante 
    
    Errori 6 - Posaggio non corretto
    Errore 7 - Errori Multipli 
    
    """
    
    status = 1
    err_string  = None
    err_counter = 0
    
    time_now = calculate_time()
    
    if yolo is not None and vbom is not None: yolo, vbom = rename_symbols(yolo), rename_symbols(vbom)
    
    try:
        if yolo is not None and yolo != []: 
                
            if frontal_check == False: 
                
                err_string = f"Errore 5 - Componente Pulsante Sbagliato  - {time_now}"
                status = 5
                
            elif frontal_check == True:
                
                #Matrici uguali ottimo
                if vbom == yolo:
                    
                    if is_alligned == True:
                    
                        err_string = (f"No Error - {time_now}\n")
                        status = -1
                        
                    elif is_alligned == False:
                            
                        err_string = f"Errore 11 - Simboli non allineati - {time_now}"
                        status = 11
                
                #Matrici diverse, capiamo perché
                elif vbom != yolo:
                                
                    #Vado colonna per colonna 
                    for column_index, column in enumerate(vbom):
                        
                        if len(vbom) == len(yolo):
                            
                            if vbom[column_index] != yolo[column_index]:
                                
                                wrong_component = class_vbom[column_index][0].componente
                                
                                if   wrong_component in ["attacco_leva"]:                                 status = 1
                                elif wrong_component in ["ghiera_1", "ghiera_2", "ghiera_3"]:             status = 2
                                elif wrong_component in ["distanziale_ghiera_1", "distanziale_ghiera_2"]: status = 3
                                elif wrong_component in ["semipomello"]:                                  status = 4
                        
                                err_string = f"Errore {status} - Componente: {wrong_component} Errato - {time_now}"
                                err_counter += 1
                                
                        elif len(vbom) < len(yolo):
                            
                            err_string = f"Errore 8 - Presenti simboli non previsti - {time_now}"
                            status = 8
                            break
                            
                        elif len(vbom) > len(yolo):
                            
                            err_string = f"Errore 9 - Assenti simboli previsti - {time_now}"
                            status = 9
                            break
                            
                if err_counter > 1: 
                    
                    err_string = f"Errore 7 - Molteplici errori Presenti, controllare Devio - {time_now}"
                    status = 7
                                     
        elif yolo is None or yolo == []: 
            
            err_string = f"Errore 7 - Nessun simbolo trovato - {time_now}"
            status = 7

        
        if status == -1: good_counter += 1
        else: bad_counter += 1
        
        if status_vector == None: status_vector = [status]
        else: status_vector.append(status)
        
        success_rate = float(good_counter/(good_counter + bad_counter))
        
        a = f"Actual Succes Rate: {float(success_rate)}, Good Counter: {good_counter}, Bad Counter: {bad_counter}"
        print(colored(a, "light_yellow"))
            
                               
    except TypeError:
        
        print("\nErrore di tipo sconosciuto, potrebbe essere passeggero\n")
                    
    return status, err_string, success_rate, good_counter, bad_counter, status_vector

#confronta le matrici
def matrices_match(class_vbom, vbom, yolo, frontal_check, fixture_check, frontal_yolo, frontal_vbom, good_counter, bad_counter):

    status = 1
    err_string  = None
    err_counter = 0
    
    time_now = calculate_time()
    
    if yolo is not None and vbom is not None: yolo, vbom = rename_symbols(yolo), rename_symbols(vbom)
    
    try:
        
        if fixture_check == False:
            

            
            err_string = f"Errore 6 - Posaggio non corretto"
            status = 6
            
        elif fixture_check == True: 
            
            if frontal_check == False: 
                
                err_string = f"Errore 1 - Componente Pulsante Sbagliato, {frontal_yolo} invece di {frontal_vbom} - {time_now}"
                
                status = 1
                
            elif frontal_check == True:
                
                if yolo is not None: 
                    
                    #Matrici uguali ottimo
                    if vbom == yolo:
                        
                        
                        
                        err_string = (f"No Error - {time_now} ")
                        status = -1
                        
                    #Matrici diverse, capiamo perché
                    elif vbom != yolo:
                        
                        #matrici con stesso numero di componenti
                        if len(vbom) == len(yolo):
                            
                            #Vado colonna per colonna 
                            for column_index, column in enumerate(vbom):
                                
                                #Se la colonne non coincidono ma la stessa sequenza di simboli si trova nella vbom
                                if vbom[column_index] != yolo[column_index] and yolo[column_index] in vbom:
                                        
                                    true_index = next((k for k, v in enumerate(yolo) if v == vbom[column_index]), None)
                            
                                    #metto [0] perchè i singoli elementi sono oggetti e hanno la proprietà componente
                                    err_string = f"Errore 2 - Componente: {class_vbom[column_index][0].componente} invertito con il componente {class_vbom[true_index][0].componente} - {time_now}"
                                    err_counter += 1
                                    status = 2
                                
                                elif vbom[column_index] != yolo[column_index] and yolo[column_index] not in vbom:
                                    
                                    err_string = f"Errore 1 - Componente: {class_vbom[column_index][0].componente} Errato - {time_now}"
                                    err_counter += 1
                                    status = 1

                        #Hanno dimensione diverse, in più o in meno
                        elif len(vbom) != len(yolo):
                            
                            for index,row in enumerate(vbom):
                                
                                if row not in yolo:
                                    
                                    err_string = f"Errore 3 - Componente: {class_vbom[index][0].componente} mancante o incompleto - {time_now}"
                                    err_counter += 1
                                    status = 3
                                    
                                    
                            for index,row in enumerate(yolo):
                                
                                if row not in vbom: 
                                    
                                    err_string = f"Errore 4 - Simboli {yolo[index]} superflui trovati"
                                    err_counter += 1
                                    status = 3
                                      
                        if float(err_counter/len(vbom)) > float(0.5): #Quando la percentuale di errore supera il 50%
                            
                            err_string = f"Errore 4 - Molteplici errori di presenti: {err_counter}, controllare Devio - {time_now}"
                            status = 4
                                     
                elif yolo is None: 
                    
                    err_string = f"Errore 4 - Nessun simbolo trovato - {time_now}"
                    status = 4
        
        if status == -1: good_counter += 1
        else: bad_counter += 1
        
        success_rate = float(good_counter/(good_counter + bad_counter))
        
        a = f"Actual Succes Rate: {float(success_rate)}, Good Counter: {good_counter}, Bad Counter: {bad_counter}"
        print(colored(a, "light_yellow"))
            
                               
    except TypeError:
        
        print("\nErrore di tipo sconosciuto, potrebbe essere passeggero\n")
    
    
        
                    
    return status, err_string, success_rate, good_counter, bad_counter
    
def packToSend(codice_devio, test_number, status, time, command, fixture):
    
    time = datetime.now()
    
    day = time.day
    month = time.month
    year = time.year
    hour = time.hour
    minute = time.minute
    second = time.second

    """
    PER QUANTO VIGNONI VORRA' MANDARLO
    stringa = time
    parti = stringa.split('/')

    anno = int(parti[0])
    mese = int(parti[1])
    giorno_ora = parti[2].split()

    giorno = int(giorno_ora[0])
    ora_min_sec = giorno_ora[1].split(':')

    ora = int(ora_min_sec[0])
    minuto = int(ora_min_sec[1])
    secondo = int(ora_min_sec[2])

    """
    if fixture != 0: fixture_code = fixture[1]
    else: fixture_code = fixture
    
    formato = str('12s10h')
    
    Struct = struct.pack(formato, codice_devio.encode('utf-8'), test_number, status, day, month, year, hour, minute, second, command, int(fixture_code) )
    
    return Struct

def pollingTera(tera12): 
    
    formato = str('2s')
    
    message = "C6"
    
    Struct = struct.pack(formato, message.encode('utf-8'))
    
    tera12.sendall(Struct)
        
def showWindows(last_status, status, current_devio):
    
    img = None
    
    #trigger generico
    if status != last_status:
        
        # print(f"{last_status}, {status}, {current_devio}, {last_devio}")
        
        # if current_devio != last_devio: cv2.destroyWindow(f"{last_devio} v-bom DOES NOT match.")
        if last_status is not None: cv2.destroyWindow('Status')
        
        #se sto riconoscendo il devio corretto
        if status ==  -1 : img = getImages(current_devio, error = False)
            
        elif (status == 1 or status == 2 or status == 3) and (last_status == None or last_status == -1): img = getImages(current_devio, error = True)
            
        #Window Dimensions and position 
        # cv2.resizeWindow(img, 700, 450)
        if img is not None: 
            
            img = cv2.imshow('Status', img)
            cv2.moveWindow(img, 750, 750)
                                
#controlla se il devio inserito è nella vbom
def checkDevioAvailability(devio_name, vbom_path):
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value == devio_name: 
                return True
    return False
    
def getImages(devio_name, error): 
    
    if error == False: 
        
        # Carica l'immagine
        img = cv2.imread('/home/ae/VBOM/popup_images/VBOM_match_base.jpg')
        # Verifica se l'immagine è stata caricata correttamente
        if img is None:
            
            print(f"Errore nel caricamento dell'immagine dal percorso: '/home/onano/Devio/VBOM_match_base.jpg'")
        else:
            
            #immagine, testo, posizione, font, scala_font, colore testo, spessore
            cv2.putText(img, f"{devio_name}", (94, 345), cv2.FONT_HERSHEY_COMPLEX, 1, (122, 70, 29), 2)
            cv2.putText(img, "V-BOM does", (44, 375), cv2.FONT_HERSHEY_COMPLEX, 1, (122, 70, 29), 2)
            cv2.putText(img, "MATCH", (264, 375), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
        
        return img       
    
    elif error == True: 
        
        # Carica l'immagine
        img = cv2.imread('/home/ae/VBOM/popup_images/VBOM_NO_match_base.jpg')

        # Verifica se l'immagine è stata caricata correttamente
        if img is None:
            
            print(f"Errore nel caricamento dell'immagine dal percorso: '/home/onano/Devio/popup_images/VBOM_NO_match_base.jpg'")
        else:
            
            #immagine, testo, posizione, font, scala_font, colore testo (BGR), spessore
            cv2.putText(img, f"{devio_name}", (104, 345), cv2.FONT_HERSHEY_COMPLEX, 1, (122, 70, 29), 2)
            cv2.putText(img, "V-BOM does", (44, 375), cv2.FONT_HERSHEY_COMPLEX, 1, (122, 70, 29), 2)
            cv2.putText(img, "NOT MATCH", (264, 375), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
        
        return img      
        
def getVBOMlist(vbom_path):
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
                
            if cell.value is not None: print(f"Codice Devio: {cell.value}")
            else: exit()
            
def setFocus(webcam, focus0_value, focus2_value, focus4_value, focus6_value):
    
    webcams = list(webcam)
    
    for webcam in webcams:
        
        if   webcam == "0":   focus_value = focus0_value
        elif webcam == "2":   focus_value = focus2_value
        elif webcam == "4":   focus_value = focus4_value
        elif webcam == "6":   focus_value = focus6_value
        
        #############CAM 1#################
        
        if focus_value is not None:
            
            #Elimino Autofocus e setto il focus a 80
            autofocus  = sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", "--get-ctrl=focus_auto"], capture_output=True, text=True)
            if autofocus.stdout == "focus_auto: 1\n": sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", "--set-ctrl=focus_auto=0"])
            sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", f"--set-ctrl=focus_absolute={focus_value}"])

            autofocus1  = sp.run(["v4l2-ctl",f"-d/dev/video{webcam}", "--get-ctrl=focus_auto"], capture_output=True, text=True)
            focus1_value = sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", "--get-ctrl=focus_absolute"], capture_output=True, text=True)

            print(colored(f"\nStato Camera {webcam}\n", "light_yellow"), 
                f"{autofocus1.stdout}",
                f"{focus1_value.stdout}")
            
        else: 
            
            a = f"\nNessun valore di focus della telecamera {webcam} inseriro, non procedo al settaggio del Focus\n"
            print(colored(a, "red")) 
    
def checkCam():
    
    video_available = []
    pattern = r'/dev/video([02468]|10)'

    command = "v4l2-ctl --list-device"
    result = sp.run(command, shell=True, stdout=sp.PIPE, text=True)
    output_lines = result.stdout.split('\n')
     
    for line in output_lines:
    
        match = re.search(pattern, line)
        if match: video_available.append(line[11])
        
    return video_available

def closeCam():
    
    video_available = checkCam()
        
    for i in video_available:
    
        autofocus  = sp.run(["v4l2-ctl", f"-d/dev/video{i}", "--get-ctrl=focus_auto"], capture_output=True, text=True)
        autoexposure  = sp.run(["v4l2-ctl", f"-d/dev/video{i}", "--get-ctrl=exposure_auto"], capture_output=True, text=True)
       
        if autofocus.stdout == "focus_auto: 0\n": sp.run([f"v4l2-ctl", f"-d/dev/video{i}", "--set-ctrl=focus_auto=1"])        
        if autoexposure.stdout == "exposure_auto: 1\n": sp.run(["v4l2-ctl", f"-d/dev/video{i}", "--set-ctrl=exposure_auto=3"])

        print(colored(f"Ripristino i parametri della camera {i}", "yellow"))

def findResultsFromString(string, folder_path):
    
    pics_results = []

    # Verifica se la cartella esiste
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' doesn't exist")
        return None

    # Elenco dei file nella cartella
    pics_name = os.listdir(folder_path)

    # Itera attraverso i file nella cartella
    for pic_name in pics_name:
        
        # Verifica se il file è un'immagine
        if pic_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
            
            # Verifica se la stringa desiderata è presente nel nome del file
            if string in pic_name:
                pics_results.append(pic_name)

    return pics_results

def checkLastResults(folder_path):
    
    # Verifica se la cartella esiste
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' doesn't exist")
        return None

    # Elenco dei file nella cartella
    pics_name = os.listdir(folder_path)

    # Itera attraverso i file nella cartella
    pic_name = [f for f in pics_name if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
        
    if not pic_name:
        print(f"The aren't results in folder '{folder_path}'")
        return None

    last_pic = max(pic_name, key=lambda f: os.path.getctime(os.path.join(folder_path, f)))

    return last_pic
      
def right_webcam_check(numero_devio, vbom_path):
    
    vbom_webcam, real_webcam = None, None    
    
    #################### LATO WEBCAM ########################################
    
    #Controllo se il numero di telecamere della VBOM corrisponde a quelle effettivamente collegate
    command = "v4l2-ctl --list-device"
    result = sp.run(command, shell=True, stdout=sp.PIPE, text=True)
    output_lines = result.stdout.split('\n')
    
    # Conta quante volte appare la stringa "HD Pro Webcam C920 o Trust Full HD Webcam"
    real_webcam = sum(1 for line in output_lines if "HD Pro Webcam C920" in line)
    real_webcam = real_webcam + sum(1 for line in output_lines if "Trust Full HD Webcam: Trust Ful" in line)
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value == numero_devio:  
                
                for i in range(1,15): 
                    
                    if sheet.cell(row=1, column = i).value == "n_webcam": webcam_col = i
                    
                vbom_webcam = sheet.cell(row=cell.row, column= webcam_col).value
                    
    if vbom_webcam == real_webcam: webcam_check = True
    else: webcam_check = False 

    # print(f"REAL WEBCAM: {real_webcam}|VBOM WEBCAM: {vbom_webcam} - Check: {check} ")

    return webcam_check, real_webcam, vbom_webcam,

def fixture_check(numero_devio, vbom_path):
    
    vbom_fixture = None
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    
    ##################### LATO POSAGGIO    ##################
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value == numero_devio:  
                
                for i in range(1,15): 
                    
                    if sheet.cell(row=1, column = i).value == "posaggio": fixture_col = i
                    
                vbom_fixture = sheet.cell(row=cell.row, column= fixture_col).value

     
    return vbom_fixture
   
def setExposure(webcam, exposure):
    
    autoexposure  = sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", "--get-ctrl=exposure_auto"], capture_output=True, text=True)
    if autoexposure.stdout == "exposure_auto: 3\n": sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", "--set-ctrl=exposure_auto=1"])
    sp.run(["v4l2-ctl", f"-d/dev/video{webcam}", f"--set-ctrl=exposure_absolute={exposure}"])
    
    
    a = f"Exposure {webcam} set to {exposure}"
    print(colored(a, "light_yellow"))

def setAllExposure(exposure0, exposure2, exposure4, exposure6):
        
    if exposure0 is not None: setExposure(0, exposure0)
    if exposure2 is not None: setExposure(2, exposure2)
    if exposure4 is not None: setExposure(4, exposure4)
    if exposure6 is not None: setExposure(6, exposure6)
    
def takeCam(current_devio, vbom_path): 
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    webcam = None
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            
            if cell.value == current_devio:  
                
                webcam = str(sheet.cell(row=cell.row, column= 11).value)
                break
                
    return webcam

def takeFrontalFocus(current_devio, vbom_path):
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["modelli_devio"]
    focus = None
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            
            if cell.value == current_devio:  
                
                focus = str(sheet.cell(row=cell.row, column= 12).value)
                break
                
    return focus

def takeExposure(current_devio, vbom_path):
    
    if current_devio == "None" or current_devio == None:
        
        print(colored("ATTENZIONE, IL DEVIO NON E' STATO TROVATO", "red"))
        
    exposure_0 = exposure_2 = exposure_4 = exposure_6 = None
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["exposure"]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            
            if cell.value == current_devio:  
                
                exposure_0 = str(sheet.cell(row=cell.row, column= 2).value)
                exposure_2 = str(sheet.cell(row=cell.row, column= 3).value)
                exposure_4 = str(sheet.cell(row=cell.row, column= 4).value)
                exposure_6 = str(sheet.cell(row=cell.row, column= 5).value)

                break
                
    return exposure_0, exposure_2, exposure_4, exposure_6

def writeExposure(current_devio, exposure_0, exposure_2, exposure_4, exposure_6, vbom_path):
    
    # Carica il foglio Excel
    wb = load_workbook(vbom_path)
    sheet = wb["exposure"]
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row: 
            
            if cell.value == current_devio:  
                
                if exposure_0 is not None and exposure_0 != 'None': 
                    
                    a = f"Scrivo nella VBOM l'exposure: {exposure_0} per la telecamera 0"
                    print(colored(a, "grey", "on_yellow"))
                    sheet.cell(row=cell.row, column= 2).value = str(round(int(exposure_0)))
                    
                if exposure_2 is not None and exposure_2 != 'None': 
                    
                    a = f"Scrivo nella VBOM l'exposure: {exposure_2} per la telecamera 2"
                    print(colored(a, "grey", "on_yellow"))
                    sheet.cell(row=cell.row, column= 3).value = str(round(int(exposure_2)))
                    
                if exposure_4 is not None and exposure_4 != 'None': 
                    
                    a = f"Scrivo nella VBOM l'exposure: {exposure_4} per la telecamera 4"
                    print(colored(a, "grey", "on_yellow"))
                    sheet.cell(row=cell.row, column= 4).value = str(round(int(exposure_4)))
                    
                if exposure_6 is not None and exposure_6 != 'None': 
                    
                    a = f"Scrivo nella VBOM l'exposure: {exposure_6} per la telecamera 6"
                    print(colored(a, "grey", "on_yellow"))
                    sheet.cell(row=cell.row, column= 5).value = str(round(int(exposure_6)))
                
                break
                
    wb.save(vbom_path)

def checkCamerafromUsb():
    
      
    # Esegui il comando e ottieni l'output
    command = "v4l2-ctl --list-device"
    result = sp.run(command, shell=True, stdout=sp.PIPE, text=True)
    
    pattern = r'Trust Full HD Webcam: Trust Ful (usb-3610000.xhci-2.1.([1234]|5)):'
    pattern1 = r'/dev/video([02468]|10)'
    
    
    # print(result.stdout)
    
    
    output_lines = result.stdout.split('\n')

    a = 'Trust Full HD Webcam: Trust Ful (usb-3610000.xhci-2.1.1):'
    b = 'Trust Full HD Webcam: Trust Ful (usb-3610000.xhci-2.1.2):'
    c = 'Trust Full HD Webcam: Trust Ful (usb-3610000.xhci-2.1.3):'
    d = 'Trust Full HD Webcam: Trust Ful (usb-3610000.xhci-2.1.4):'
  
    for line in output_lines:
        
        if a == line: print("True")
        if b == line: print("True") 
        if c == line: print("True") 
        if d == line: print("True")
        
        match = re.search(pattern, line)
        match1 = re.search(pattern1, line)
        
        if match: print("Match")
        if match1: print("Match1")

        # print(output_lines)
        
def matrix_merge(devio, right_matrix, frontal_matrix, left_matrix, debug):
    
    pricipal_matrix = copy.deepcopy(frontal_matrix)
    
    if pricipal_matrix is None: 
        
        if debug == True: print(colored(f"\nLa matrice frontale non è stata trovata per il devio", "red"))
        
        return None
    
    frontal_stripes_counter = lateral_stripes_counter = right_stripes_counter = 0
    
    merged_column = []
    next_column = False
        
    if right_matrix is not None:
            
        if debug == True: print(colored(f"\nLa matrice destra è stata trovata"))
        
        #itero per ogni colonna della matrice laterale per appendere i valori in quella superiore
        for right_i, right_column in enumerate(right_matrix):
            
            next_column = False
            
            #itero per ogni elemento delle colonne della matrice laterale
            for right_j, right_symbol in enumerate(right_column):
                
                #salto la colonna se il flag è alzato 
                if next_column == True: 
                    
                    if debug == True: print("Passo alla nuova colonna")
                    continue
                
                if right_symbol == "3_narrow_stripes" or right_symbol == "2_bold_stripes": 
                    
                    right_stripes_counter += 1
                    if debug == True: print(f"\nStripes Destra n'{right_stripes_counter} passata")
                    
                #Se l'indice della colonna della matrice laterale supera il numero delle colonne della matrice frontale
                if right_i < len(pricipal_matrix):
                    
                    a = f"\nDebug: [Colonna n'{right_i}|{right_column}, Simbolo n'{right_j}|{right_symbol} della matrice destra\n"
                    if debug == True: print(colored(a, "yellow"))
                    
                    frontal_stripes_counter = 0 
                    
                    if debug == True: print(f"Rientro nella matrice frontale, n'stripes = {frontal_stripes_counter}")
                    
                    #itero, partendo da singolo simbolo dalla matrice lateral, all'interno della matrice frontale
                    for frontal_i, frontal_column in enumerate(pricipal_matrix):
                        
                        if "3_narrow_stripes" in frontal_column or "2_bold_stripes" in frontal_column: 
                            
                            frontal_stripes_counter += 1
                            if debug == True: print(f"\nStripes Frontale n'{frontal_stripes_counter} passata")
                                
                        #se il simbolo trovato nella matrice sinistra si trova nella colonna della matrice centrale, 
                        if right_symbol in frontal_column  and frontal_stripes_counter == right_stripes_counter and frontal_column[-1] != "1_circle" and right_column[-1] != "1_circle":  #and frontal_column[-1] == left_symbol   ||| len(frontal_column) >= 2
                            
                            if debug == True: print(f"\nIl simbolo {right_symbol} è presente nella riga {frontal_i} della matrice frontale ed è dopo la stripes frontale {frontal_stripes_counter} e destra {right_stripes_counter}")
                            
                            #Una volta che il simbolo appartenente alla matrice laterale è l'ultimo estremo di una colonne della matrice centrale, appendi le colonne                            
                            for item in right_matrix[right_i] + pricipal_matrix[frontal_i]: 
                                
                                #Se è la prima volta che lo vedo, lo inserico 
                                if item not in merged_column: 
                                    
                                    a = f"\nAggiungo {item} alla colonna, considerando che {item} non è in {merged_column}"
                                    merged_column.append(item)
                                    if debug == True: print(colored(a, "green"))
                                
                                #se invece è già presente, passo oltre
                                elif item in merged_column: 
                                    
                                    a = f"\n{item} è già presente nella colonna {merged_column}, quindi non lo appendo"
                                    if debug == True: print(colored(a, "red"))
                                    
                                    
                                a = f"\nLa colonna attuale è: {merged_column}"
                                if debug == True: print(colored(a, "yellow"))
                        
                            pricipal_matrix[frontal_i] = merged_column
                            
                            if debug == True: 
                                print(f"\n\nLa matrice attualmente è:")
                                for i in pricipal_matrix: 
                                    if debug == True: print(f"{i}")
                            
                            merged_column = []
                            next_column = True
                        
                        elif right_symbol not in frontal_column:
                            
                            if debug == True: print(f"\nIl simbolo {right_symbol} non è presente nella riga {frontal_i} della matrice frontale")
                            pass
                        
                        elif frontal_stripes_counter != right_stripes_counter:
                            
                            if debug == True: print(f"Il simbolo corrisponde a una zona diversa, {right_symbol} è dopo la {right_stripes_counter} stripes, invece la colonna {frontal_column} è dopo la {frontal_stripes_counter} stripes")
    
                        elif frontal_column[-1] == "1_circle":
                            
                            if debug == True: print(f"Il simbolo 1_circle è presente nella colonna, ma è l'ultimo, e se è l'ultimo non lo uso per")

                        elif right_column[-1] == "1_circle":
                            
                            if debug == True: print(f"Il simbolo 1_circle è presente nella colonna, ma è l'ultimo, e se è l'ultimo non lo uso per")
                            
    frontal_stripes_counter = lateral_stripes_counter = right_stripes_counter = 0     
     
    merged_column = []
    next_column = False
    
    if left_matrix is not None:
        
        if debug == True: print(colored(f"\nLa matrice sinistra è stata trovata"))
        
        #itero per ogni colonna della matrice laterale per appendere i valori in quella superiore
        for left_i, left_column in enumerate(left_matrix):
            
            next_column = False
            
            #itero per ogni elemento delle colonne della matrice laterale
            for left_j, left_symbol in enumerate(left_column):
                
                #salto la colonna se il flag è alzato 
                if next_column == True: 
                    
                    if debug == True: print("Passo alla nuova colonna")
                    continue
                
                if left_symbol == "3_narrow_stripes" or left_symbol == "2_bold_stripes": 
                    
                    lateral_stripes_counter += 1
                    if debug == True: print(f"\nStripes Laterale n'{lateral_stripes_counter} passata")
                
                #Se l'indice della colonna della matrice laterale supera il numero delle colonne della matrice frontale
                if left_i < len(pricipal_matrix):
                    
                    a = f"\nDebug: [Colonna n'{left_i}|{left_column}, Simbolo n'{left_j}|{left_symbol} della matrice laterale\n"
                    if debug == True: print(colored(a, "yellow"))
                    
                    frontal_stripes_counter = 0
                    if debug == True: print(f"Rientro nella matrice frontale, n'stripes = {frontal_stripes_counter}")
                    
                    #itero, partendo da singolo simbolo dalla matrice lateral, all'interno della matrice frontale
                    for frontal_i, frontal_column in enumerate(pricipal_matrix):
                        
                        if "3_narrow_stripes" in frontal_column or "2_bold_stripes" in frontal_column: 
                            
                            frontal_stripes_counter += 1
                            if debug == True: print(f"\nStripes Frontale n'{frontal_stripes_counter} passata")
                            
                        #se il simbolo trovato nella matrice sinistra si trova nella colonna della matrice centrale, 
                        if left_symbol in frontal_column  and frontal_stripes_counter == lateral_stripes_counter and frontal_column[-1] != "1_circle" and left_column[-1] != "1_circle":  #and frontal_column[-1] == left_symbol   ||| len(frontal_column) >= 2
                            
                            if debug == True: print(f"\nIl simbolo {left_symbol} è presente nella riga {frontal_i} della matrice frontale ed è dopo la stripes frontale {frontal_stripes_counter} e laterale {lateral_stripes_counter}")
                            
                            #Una volta che il simbolo appartenente alla matrice laterale è l'ultimo estremo di una colonne della matrice centrale, appendi le colonne                            
                            for item in pricipal_matrix[frontal_i] + left_matrix[left_i]: 
                                
                                #Se è la prima volta che lo vedo, lo inserico 
                                if item not in merged_column: 
                                    
                                    a = f"\nAggiungo {item} alla colonna, considerando che {item} non è in {merged_column}"
                                    merged_column.append(item)
                                    if debug == True: print(colored(a, "green"))
                                
                                #se invece è già presente, passo oltre
                                elif item in merged_column: 
                                    
                                    a = f"\n{item} è già presente nella colonna {merged_column}, quindi non lo appendo"
                                    if debug == True: print(colored(a, "red"))
                                    
                                a = f"\nLa colonna attuale è: {merged_column}"
                                if debug == True: print(colored(a, "yellow"))
                        
                            pricipal_matrix[frontal_i] = merged_column
                            
                            if debug == True: 
                                print(f"\n\nLa matrice attualmente è:")
                                for i in pricipal_matrix: 
                                    if debug == True: print(f"{i}")
                            
                            merged_column = []
                            next_column = True
                        
                        elif left_symbol not in frontal_column:
                            
                            if debug == True: print(f"\nIl simbolo {left_symbol} non è presente nella riga {frontal_i} della matrice frontale")
                            pass
                        
                        elif frontal_stripes_counter != lateral_stripes_counter:
                            
                            if debug == True: print(f"Il simbolo corrisponde a una zona diversa, {left_symbol} è dopo la {lateral_stripes_counter} stripes, invece la colonna {frontal_column} è dopo la {frontal_stripes_counter} stripes")
    
                        elif frontal_column[-1] == "1_circle":
                            
                            if debug == True: print(f"Il simbolo 1_circle è presente nella colonna, ma è l'ultimo, e se è l'ultimo non lo uso per")

                        elif left_column[-1] == "1_circle":
                                
                                if debug == True: print(f"Il simbolo 1_circle è presente nella colonna, ma è l'ultimo, e se è l'ultimo non lo uso per")
    
    return pricipal_matrix
    
def define_two_max_exposure(vectors, gain, minimum_fixture):

    vectors = sorted(vectors, key=lambda x: x[1]) #metto in ordine i valori    
    sorted_vector = sorted(vectors, key=lambda x: x[0], reverse = True)
    
    if minimum_fixture is not None: sorted_vector = filtring(sorted_vector, minimum_fixture)
    
    maxtwovalue = sorted_vector[0:2]
    
    # for i in sorted_vector: print(i)
    
    #Estraggo le probabilità e le esposizioni
    prob_a = maxtwovalue[1][0]
    prob_b = maxtwovalue[0][0]
    
    exposure_a = maxtwovalue[1][1]
    exposure_b = maxtwovalue[0][1]
    
    # print(f" b expsoure/prob: {exposure_b}/{prob_b}, a exposure/prob: {exposure_a}/{prob_a}")
    
    #Scremo nel caso in cui le due esposizioni massime sono due massimi relativi distanti 
    if abs(exposure_b - exposure_a) > gain: 
        
        print("Massimi relativi distanti, cerchiamo l'esposizione migliore nell'intorno del Massimo Assoluto")
        
        for i, vector in enumerate(vectors): 
            
            if vector[1] == exposure_b: b_index = i #quando trovo l'indice di riferimento di exposure_b
        # print(f"b_index: {b_index}")
            
        # print(len(vectors), b_index)
        
        #Se siamo alla fine del vettore
        if b_index + 1 == len(vectors):
            
            exposure_a = vectors[b_index - 1][1]
            prob_a = vectors[b_index - 1][0]
            
        #se siamo all'inizio del vettore
        elif b_index == 0: 
            
            exposure_a = vectors[b_index+1][1]
            prob_a = vectors[b_index + 1][0]
        
        #se siamo a metà
        else:
            
            right_b_exoposure = vectors[b_index + 1][1]
            right_b_prob = vectors[b_index + 1][0]
            
            left_b_exposure = vectors[b_index - 1][1] 
            left_b_prob = vectors[b_index - 1][0]
            
            #cerco il migliore 
            if right_b_prob > left_b_prob: 
                
                exposure_a = right_b_exoposure
                prob_a = right_b_prob
                
            elif right_b_prob < left_b_prob: 
                
                exposure_a = left_b_exposure  
                prob_a = left_b_prob
                
            
        # print(f"Central Exposure/Prob: {exposure_b}/{prob_b}\nRight Exposure/Prob: {right_b_exoposure}/{right_b_prob}\nLeft Exposure/Prob: {left_b_exposure}/{left_b_prob}")
        if exposure_a > exposure_b: exposure_b, prob_b, exposure_a, prob_a = exposure_a, prob_a, exposure_b, prob_b
        
    return exposure_b, exposure_a, prob_b, prob_a, maxtwovalue

def rename_duplicates(coordinate_list):
    name_count = {}
    renamed_list = []

    for item in coordinate_list:
        name = item[0]
        if name in name_count:
            name_count[name] += 1
            new_name = f"{name}_{name_count[name]}"
        else:
            name_count[name] = 1
            new_name = name
        renamed_list.append([new_name, item[2], item[3]])

    return renamed_list

def rename_symbols(coordinate_list):
    name_count = {}
    renamed_list = []

    for item in coordinate_list:
        # Prendere il primo elemento della lista come nome
        name = item[0]
        if name in name_count:
            name_count[name] += 1
            new_name = f"{name}_{name_count[name]}"
        else:
            name_count[name] = 1
            new_name = name
        # Creare una nuova lista con il nuovo nome e gli altri elementi originali
        renamed_list.append([new_name] + item[1:])

    return renamed_list

def checkAllignment(current_devio, threshold, coordinate_list, background_img, max_slope, debug): 
    
    is_alligned = False
    point1 = point2 = point3 = point4 = None
    points = []
    
    coordinate_list.sort(key=lambda x: (x[2], -x[3]))  # Ordina per X crescente e Y decrescente
    item = rename_duplicates(coordinate_list)
    
    coordinates = {item[0]: (item[1], 1 - item[2]) for item in rename_duplicates(coordinate_list)}

   
    if current_devio == "01-1430-0000": #dot dot e circle
        
        if all(key in coordinates for key in ["0_dot", "0_dot_2", "1_circle", "15_rear_wipers_02"]):
            
            point1 = coordinates["0_dot"]
            point2 = coordinates["0_dot_2"]
            point3 = coordinates["1_circle"]
            point4 = coordinates["15_rear_wipers_02"]
        
    elif current_devio == "01-1465-0000": #circle dot circle
        
        if all(key in coordinates for key in ["1_circle", "0_dot", "1_circle_2"]):
            
            point1 = coordinates["1_circle"]
            point2 = coordinates["0_dot"]
            point3 = coordinates["1_circle_2"]
        
    elif current_devio == "01-1466-0000" or current_devio == "01-1467-0000": #dot circle e turn signal 
        
        if all(key in coordinates for key in ["0_dot", "6_turn_sig", "1_circle"]):
            
            point1 = coordinates["0_dot"]
            point2 = coordinates["1_circle"]
            point3 = coordinates["6_turn_sig"]
        
        
    elif current_devio == "01-1437-0000": #circle e dot e circle
        
        if all(key in coordinates for key in ["1_circle", "0_dot", "1_circle_2"]):
            
            point1 = coordinates["1_circle"]
            point2 = coordinates["0_dot"]
            point3 = coordinates["1_circle_2"]
        
    elif current_devio == "01-1429-0000": #dot e circle turn signal 
        
        if all(key in coordinates for key in ["0_dot", "6_turn_sig", "1_circle"]):
            
            point1 = coordinates["0_dot"]
            point2 = coordinates["1_circle"]
            point3 = coordinates["6_turn_sig"]
        
        
    elif current_devio == "01-1438-0000" or current_devio == "01-1439-0000": #dot circle e turn signal 
        
        if all(key in coordinates for key in ["0_dot", "6_turn_sig", "1_circle"]):
            
            point1 = coordinates["0_dot"]
            point2 = coordinates["1_circle"]
            point3 = coordinates["6_turn_sig"]

    else: 
        print("Non posso definire l'allineamento di questo devio, non ho i dati per farlo")
        is_alligned = True 
        
    if point1 is not None and point2 is not None and point3 is not None: 

        if point4 is not None: points = [point1, point2, point3, point4]
        else: points = [point1, point2, point3]    
    
    #prendo le coordinate del primo e dell'ultimo e ne calcolo il coefficente e per una retta
    if points != []: 
        
        m = (points[-1][1] - points[0][1]) / (points[-1][0] - points[0][0])
        q = points[0][1] - m * points[0][0]
    
        #la retta ora è y = mx + q

        if (abs(m) < max_slope): 
            
            if point4 is None:
                if (abs(points[1][1] - (m * points[1][0] + q)) < threshold): is_alligned = True
            elif point4 is not None: 
                if abs(points[2][1] - (m * points[2][0] + q)) < threshold and abs(points[1][1] - (m * points[1][0] + q)) < threshold: is_alligned = True
        
        else: print(f"La retta è troppo inclinata, {abs(m)} > {max_slope}")
         
        if debug == True: print(f"Il punto {points[1][0]}, {points[1][1]} è distante di {abs(points[1][1] - (m * points[1][0] + q))} dalla retta: y = {m}x + {q}")
        if debug == True and point4 is not None: print(f"Il punto {points[2][0]}, {points[2][1]} è distante di {abs(points[2][1] - (m * points[2][0] + q))} dalla retta: y = {m}x + {q}")
            
        # Disegna la retta centrale
        img_height, img_width = background_img.shape[:2]
        x_values = np.linspace(0, 1, 500)
        y_values = m * x_values + q
        
        line_points = np.array([(int(x * img_width), int((1 - y) * img_height)) for x, y in zip(x_values, y_values)], np.int32)
        line_points_left = np.array([(int(x * img_width), int((1 - y - threshold) * img_height))  for x, y in zip(x_values, y_values)], np.int32)
        line_points_right = np.array([(int(x * img_width), int((1 - y + threshold) * img_height)) for x, y in zip(x_values, y_values)], np.int32)
        
        for line in [line_points, line_points_left, line_points_right]: cv2.polylines(background_img, [line], isClosed=False, color=(255, 0, 0), thickness=2)

        coordinate_values = list(coordinates.values()) 
        
        # Disegna i punti
        for point in coordinate_values:
            cv2.circle(background_img, (int(point[0] * img_width), int((1 - point[1]) * img_height)), 5, (0, 0, 255), -1)
    
    
    # print(f"is_alligned: {is_alligned}")
    
    return is_alligned

def remove_concetric_symbols(coordinate_list, threshold):
    
    filtered_coordinate_list = []

    # Iterate over the list of vectors
    for i in range(len(coordinate_list)):
        keep = True
        for j in range(len(coordinate_list)):
            if i != j:
                if abs(coordinate_list[i][2] -  coordinate_list[j][2]) <= threshold and abs(coordinate_list[i][3] - coordinate_list[j][3]) <= threshold:
                    if coordinate_list[i][1] < coordinate_list[j][1]:
                        keep = False
                        break
        if keep:
            filtered_coordinate_list.append(coordinate_list[i])
            
        else: 
            
            print(f"Ho rimosso {coordinate_list[i][0]} con conf: {coordinate_list[i][1]} perché nello stesso posto di {coordinate_list[j][0]} con conf {coordinate_list[j][1]}")

    return filtered_coordinate_list

def more_frequent(vector):
    
    counter = Counter(vector)
    more_frequent_value = counter.most_common(1)[0][0]
    return more_frequent_value

def takeandsetAllExposure(numero_devio, vbom_path):
    
    exposure_0, exposure_2, exposure_4, exposure_6 = takeExposure(numero_devio, vbom_path)
        
    setExposure("0", exposure_0)
    setExposure("2", exposure_2)
    setExposure("4", exposure_4)
    setExposure("6", exposure_6)
    
def filtring(vector_list, threshold):
    
    filtered_vector_list = [sottolista for sottolista in vector_list if sottolista[3] >= 4]
    return filtered_vector_list

def extract_confidence(data):
    
    # Usa una regex per trovare il valore di 'confidence'
    match = re.search(r'confidence:\s*([0-9]*\.?[0-9]+)', data)
    if match:
        return float(match.group(1))
    else:
        print("Confidence value not found")
        return None

def extract_pn_esito_time(data):
    
    part_number = re.search(r'Part Number:\s*(.*)', data).group(1)
    esito = re.search(r'esito:\s*(-?\d+)', data).group(1)
    time = re.search(r'time:\s*(.*)', data).group(1)
    
    return part_number, esito, time

def newcheckLastResults(folder_path, debug):
    
    periodicity = setImagesPeriodicity(folder_path)
        
    confidence_0 = confidence_2 = confidence_4 = confidence_6 = float(0)
    part_number = esito = time =  None

    # Verifica se la cartella esiste
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' doesn't exist")
        return None

    # Elenco dei file nella cartella
    pics_name = os.listdir(folder_path)

    # Itera attraverso i file nella cartella
    pic_name = [f for f in pics_name if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
        
    if not pic_name:
        print(f"The aren't results in folder '{folder_path}'")
        return None

   # Ordina le immagini per data di creazione in ordine decrescente
    sorted_pics = sorted(pic_name, key=lambda f: os.path.getctime(os.path.join(folder_path, f)), reverse=True)
    
    if debug == 1: 
        print(f"\nLAST TEN PICS\n")
        for i, pic in enumerate(sorted_pics[:10]):
            
            creation_time = os.path.getctime(os.path.join(folder_path, pic))
            creation_time_str = datetime.fromtimestamp(creation_time).strftime('%Y-%m-%d %H:%M:%S')
            print(f"{pic.strip()} - Created on: {creation_time_str}")
            print('-'*40)
        
    # Prende le prime n immagini (o meno se non ci sono abbastanza immagini)
    if periodicity is not None: last_pics = sorted_pics[:periodicity*2]
        
    if debug == 1: 
        print("\nLAST PICS\n")
        for last_pic in last_pics:
            print(last_pic.strip())
            print('+'*40)
    
    for pic in last_pics: 
        
        if "NAKED" not in pic: 
            
            if   "Cam: 0" in pic: confidence_0 = extract_confidence(pic)
            elif "Cam: 2" in pic: confidence_2 = extract_confidence(pic)
            elif "Cam: 4" in pic: confidence_4 = extract_confidence(pic)
            elif "Cam: 6" in pic: confidence_6 = extract_confidence(pic)
    
        part_number, esito, time = extract_pn_esito_time(pic)
                
    confidences = [confidence_0, confidence_2, confidence_4, confidence_6]

    report = createReport(confidences, part_number, esito, time)
    
    return report

def createReport(confidences, part_number, esito, time):
    
    
    webcams = [0,2,4,6]
    
    report_lines = []
    
    
    confidence_string = "confidence:"+ "".join(list(f" C{cam} {confidence:.2f}," for cam, confidence in zip(webcams, confidences)))
    
    report_lines.append(f"Part Number: {part_number}")
    report_lines.append(confidence_string) 
    report_lines.append(f"esito: {esito}")
    report_lines.append(f"time: {time}")

    report = "\n".join(report_lines)
    return report

def setImagesPeriodicity(path): 
    
    is_four_here = False 
    
    # Verifica se la cartella esiste
    if not os.path.exists(path):
        print(f"The folder '{path}' doesn't exist")
        return None

    # Elenco dei file nella cartella
    pics_name = os.listdir(path)

    # Itera attraverso i file nella cartella
    pic_name = [f for f in pics_name if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
        
    if not pic_name:
        print(f"The aren't results in folder '{path}'")
        return None

    # Ordina le immagini per data di creazione in ordine decrescente
    sorted_pics = sorted(pic_name, key=lambda f: os.path.getctime(os.path.join(path, f)), reverse=True)
    
    first_four_pics = sorted_pics[:4]
   
    for pic in first_four_pics: 
       
       if "Cam: 4" in pic: is_four_here = True
       
    if is_four_here == True: return 4
    else: return 3
    
if __name__ == '__main__':
    
    
    upper_matrix = [
        ['0_dot'],
        ['1_circle', '4_pos', '26_R_anabb'],
        ['3_narrow_stripes'],
        ['6_turn_sig']]
    
    left_matrix = [
        ['0_dot'],
        ['1_circle', '4_pos', '26_R_anabb', '36_R_auto_anabb'],
        ['3_narrow_stripes'],
        ['6_turn_sig', '29_R_abb_switch_01']]
    
    right_matrix = None 
    
    debug = False 
    
    print(upper_matrix)
        
    merged_matrix, upper_matrix = matrix_merge("01-1430-0000", right_matrix, upper_matrix, left_matrix, debug)
        
    print(upper_matrix)